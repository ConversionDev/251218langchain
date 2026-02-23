"""
매핑 결과를 사람이 보기 편한 Excel(.xlsx)로 저장.
시트명 "환경물질 매핑", 컬럼: CAS 번호, EC 번호, 배터리관련, 12개 언어별 물질명·동의어·검증.
"""

from pathlib import Path
from openpyxl.utils import get_column_letter

from .config import LANG_CODE_TO_LABEL, excel_col_name


def _excel_column_order() -> list[str]:
    """Excel 컬럼 순서."""
    cols = ["CAS 번호", "EC 번호", "배터리관련"]
    for code, label in LANG_CODE_TO_LABEL.items():
        cols.append(excel_col_name("물질명", label))
        cols.append(excel_col_name("동의어", label))
        cols.append(excel_col_name("검증", label))
    return cols


def build_rows(substances: list[dict], lang_results: dict[tuple[str, str], dict]) -> list[dict]:
    """물질 목록 + (cas, lang_code)별 결과 → Excel용 행 리스트 (CAS당 1행)."""
    rows = []
    for sub in substances:
        cas = sub.get("cas", "")
        row = {
            "CAS 번호": cas,
            "EC 번호": sub.get("ec_number", ""),
            "배터리관련": "Y" if sub.get("battery_related") else "N",
        }
        for (c, lang_code), data in lang_results.items():
            if c != cas:
                continue
            label = LANG_CODE_TO_LABEL.get(lang_code, lang_code)
            row[excel_col_name("물질명", label)] = data.get("name", "")
            syns = data.get("synonyms") or []
            row[excel_col_name("동의어", label)] = " · ".join(syns) if isinstance(syns, list) else str(syns)
            row[excel_col_name("검증", label)] = data.get("status", "")
        rows.append(row)
    return rows


def write_excel(rows: list[dict], out_path: Path, sheet_name: str = "환경물질 매핑") -> None:
    """행 리스트를 Excel로 저장. 컬럼 너비·헤더 굵게·첫 행·좌측 3열 고정."""
    import pandas as pd
    from openpyxl import load_workbook

    order = _excel_column_order()
    df = pd.DataFrame(rows)
    for c in order:
        if c not in df.columns:
            df[c] = ""
    df = df[order]
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(out_path, index=False, sheet_name=sheet_name, engine="openpyxl")
    wb = load_workbook(out_path)
    ws = wb.active
    ws.title = sheet_name
    for i, c in enumerate(order, 1):
        cell = ws.cell(1, i)
        cell.font = cell.font.copy(bold=True)
        cell.alignment = cell.alignment.copy(horizontal="center")
    widths = {"CAS 번호": 14, "EC 번호": 14, "배터리관련": 14}
    for i, col in enumerate(order, 1):
        w = widths.get(col, 22)
        if "동의어" in col:
            w = 38
        elif "검증" in col:
            w = 12
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "D2"
    wb.save(out_path)
